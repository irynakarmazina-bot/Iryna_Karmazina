"""
Читання таблиці рахунків Маерска (Excel) → список рахунків для рознесення.

Формат таблиці (Overdue Invoices, експорт Maersk):
  Product type | Invoice number | Reference - e.g. BL | Invoice type |
  Invoice date | Due date | Invoiced amount | Currency | Open amount | Currency | ...

Колонки шукаються ЗА НАЗВОЮ заголовка (а не за жорстким номером),
щоб читалка не ламалась, якщо порядок колонок трохи зміниться.
"""
from pathlib import Path

import openpyxl


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _find_header_row(ws):
    """Знайти рядок заголовків (той, де є 'invoice number' і 'reference')."""
    for r in range(1, min(ws.max_row, 20) + 1):
        cells = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        joined = " | ".join(cells)
        if "invoice number" in joined and "reference" in joined:
            return r, cells
    raise ValueError("Не знайдено рядок заголовків (немає 'Invoice number'/'Reference')")


def _col_index(headers, *needles):
    """Номер колонки (1-based), де заголовок містить будь-яке зі слів."""
    for i, h in enumerate(headers):
        if any(n in h for n in needles):
            return i + 1
    return None


def read_invoices(path: str) -> list[dict]:
    """
    Повертає список рахунків. Кожен — dict:
      bl             — номер коносамента (Reference), рядком
      invoice_number — номер рахунку Маерска (рядком)
      invoice_type   — тип (напр. DEM-INV)
      invoice_date   — дата рахунку (рядком, як у файлі)
      amount         — сума рядком (кол. 'Invoiced amount', як у файлі, кома як десятк.)
      amount_num     — та сама сума числом (для перевірок/підсумків)
      currency       — валюта (напр. USD)
    Рядок-підсумок ('Total'/порожній номер рахунку) пропускається.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header_row, headers = _find_header_row(ws)

    c_num = _col_index(headers, "invoice number")
    c_bl = _col_index(headers, "reference")
    c_type = _col_index(headers, "invoice type")
    c_date = _col_index(headers, "invoice date")
    c_amt = _col_index(headers, "invoiced amount")
    # Валюта — перша колонка 'currency' одразу після 'invoiced amount'
    c_cur = c_amt + 1 if c_amt else _col_index(headers, "currency")

    missing = [name for name, idx in (
        ("Invoice number", c_num), ("Reference/BL", c_bl),
        ("Invoiced amount", c_amt), ("Currency", c_cur),
    ) if idx is None]
    if missing:
        raise ValueError("У таблиці бракує колонок: " + ", ".join(missing))

    invoices = []
    for r in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(r, c_num).value
        bl = ws.cell(r, c_bl).value
        # Пропустити порожні рядки і рядок-підсумок (немає номера рахунку або BL)
        if num is None or str(num).strip() == "":
            continue
        if bl is None or str(bl).strip() == "":
            continue

        amt_raw = ws.cell(r, c_amt).value
        amount = str(amt_raw).strip() if amt_raw is not None else ""
        try:
            amount_num = float(amount.replace(" ", "").replace(" ", "").replace(",", "."))
        except ValueError:
            amount_num = None

        invoices.append({
            "bl": str(bl).strip(),
            "invoice_number": str(num).strip(),
            "invoice_type": str(ws.cell(r, c_type).value or "").strip() if c_type else "",
            "invoice_date": str(ws.cell(r, c_date).value or "").strip() if c_date else "",
            "amount": amount,
            "amount_num": amount_num,
            "currency": str(ws.cell(r, c_cur).value or "").strip(),
        })

    return invoices


if __name__ == "__main__":
    import sys

    rows = read_invoices(sys.argv[1])
    print(f"Знайдено рахунків: {len(rows)}")
    total = 0.0
    for inv in rows:
        print(f"  BL {inv['bl']} | рах. {inv['invoice_number']} | "
              f"{inv['invoice_type']} | {inv['amount']} {inv['currency']} "
              f"| {inv['invoice_date']}")
        if inv["amount_num"]:
            total += inv["amount_num"]
    print(f"Сума Invoiced: {total:.2f}")
