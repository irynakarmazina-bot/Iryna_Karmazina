"""Збірка «Оперативного фінансового звіту» з реальних вивантажень Експедитора."""
from datetime import date, datetime, timedelta
from collections import defaultdict
import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

U = '/root/.claude/uploads/72d9fdc9-c0c9-56db-b030-7685c1722c4c/'
TODAY = date(2026, 7, 3)
MON = TODAY - timedelta(days=TODAY.weekday())  # понеділок поточного тижня

def num(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int, float)): return float(v)
    return float(str(v).replace(' ', '').replace('\xa0', '').replace(',', '.'))

def dt(v):
    if v is None or v == '': return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

def deal_no(v):
    # «Угода 00000021 від 17.11.2025 10:19:09» → «У-21»
    s = str(v or '')
    if 'Угода' in s:
        try: return 'У-' + s.split()[1].lstrip('0')
        except Exception: return s[:20]
    return s[:20]

# ── Читання файлів ──
wb_in = load_workbook(glob.glob(U + 'c8f5e242*')[0], read_only=True, data_only=True)
inc = []
for r in list(wb_in['Sheet1'].iter_rows(values_only=True))[1:]:
    if r[2] is None: continue
    inc.append(dict(
        num=str(r[0] or ''), date=dt(r[2]), client=str(r[17] or r[4] or '?'),
        deal=deal_no(r[16]), cur=str(r[7] or ''), amount=num(r[8]), uo=num(r[9]),
        status=str(r[11] or ''), paid_date=dt(r[12]), due=dt(r[42]),
        containers=str(r[6] or ''), kasa=str(r[3] or ''),
    ))

wb_ex = load_workbook(glob.glob(U + '65f80089*')[0], read_only=True, data_only=True)
exp = []
for r in list(wb_ex['Sheet1'].iter_rows(values_only=True))[1:]:
    if r[1] is None: continue
    exp.append(dict(
        num=str(r[1] or ''), date=dt(r[2]), supplier=str(r[6] or '?'),
        deal=deal_no(r[5]), item=str(r[9] or 'Без статті'), cur=str(r[11] or ''),
        uo=num(r[13]), status=str(r[16] or ''), paid_date=dt(r[17]), due=dt(r[36]),
        advance=(str(r[7]) == 'Так'), note=str(r[19] or ''), kasa=str(r[8] or ''),
        rem=(0.0 if str(r[16]) == 'Сплачений' else (num(r[18]) if r[18] not in (None, '') else num(r[13]))),
    ))
    exp[-1]['paid_uo'] = max(0.0, exp[-1]['uo'] - exp[-1]['rem'])
if False:
    _ = dict((
    ))

wb_b = load_workbook(glob.glob(U + 'ff3100f7*')[0], read_only=True, data_only=True)
bal = []
for r in list(wb_b['Sheet1'].iter_rows(values_only=True))[3:]:
    if r[0] and str(r[0]) != 'Разом':
        bal.append(dict(acc=str(r[0]), cur=str(r[2] or ''), amount=num(r[3]), uo=num(r[5])))

wb_d = load_workbook(glob.glob(U + '20e4b397*')[0], read_only=True, data_only=True)
deals = []
for r in list(wb_d['Sheet1'].iter_rows(values_only=True))[1:]:
    if r[1] is None: continue
    route = ' → '.join([str(p) for p in (r[10], r[11], r[12]) if p])
    deals.append(dict(
        num='У-' + str(r[1]), date=dt(r[2]), client=str(r[8] or '?'),
        status=str(r[7] or ''), typ=str(r[6] or ''), route=route[:40],
        cargo=str(r[24] or ''), plan_profit=num(r[20]),
    ))

wb_ad = load_workbook(glob.glob(U + 'b61b4230*')[0], read_only=True, data_only=True)
all_deals = {}
for r in list(wb_ad['Sheet1'].iter_rows(values_only=True))[1:]:
    if r[1] is None: continue
    all_deals['У-' + str(r[1])] = dict(client=str(r[8] or '?'), status=str(r[7] or ''), date=dt(r[2]))

def client_of(x, fallback=''):
    d = all_deals.get(x['deal'])
    return d['client'] if d else (fallback or '(без угоди)')

inc_unpaid = [x for x in inc if x['status'] != 'Сплачений']
exp_unpaid = [x for x in exp if x['status'] != 'Сплачений']

# Середній термін оплати по клієнтах (з оплачених)
dso = {}
for cl in {x['client'] for x in inc}:
    ds = [(x['paid_date'] - x['date']).days for x in inc
          if x['client'] == cl and x['paid_date'] and x['date'] and x['status'] == 'Сплачений']
    if ds: dso[cl] = round(sum(ds) / len(ds))


# Очікувані дати оплати (планова дата Експедитора ненадійна — рахуємо з фактів)
import statistics
all_days = [(x['paid_date'] - x['date']).days for x in inc
            if x['paid_date'] and x['date'] and x['status'] == 'Сплачений']
MED = round(statistics.median(all_days)) if all_days else 10
dpo = {}
for sp in {x['supplier'] for x in exp}:
    ds = [(x['paid_date'] - x['date']).days for x in exp
          if x['supplier'] == sp and x['paid_date'] and x['date'] and x['status'] == 'Сплачений']
    if ds: dpo[sp] = round(sum(ds) / len(ds))
MED_EXP = round(statistics.median([(x['paid_date'] - x['date']).days for x in exp
                if x['paid_date'] and x['date'] and x['status'] == 'Сплачений'])) if exp else 10
for x in inc_unpaid:
    x['due_calc'] = x['date'] + timedelta(days=dso.get(x['client'], MED)) if x['date'] else TODAY
for x in exp_unpaid:
    x['due_calc'] = x['date'] + timedelta(days=dpo.get(x['supplier'], MED_EXP)) if x['date'] else TODAY



# ── Аналітика ──
debitorka = sum(x['uo'] for x in inc_unpaid)
kreditorka = sum(x['rem'] for x in exp_unpaid)
avansy = sum(x['rem'] for x in exp_unpaid if x['advance'])
money_uo = sum(b['uo'] for b in bal)

overdue_in = [x for x in inc_unpaid if x.get('due_calc') and x['due_calc'] < TODAY]
overdue_ex = [x for x in exp_unpaid if x.get('due_calc') and x['due_calc'] < TODAY]
res7 = sum(x['rem'] for x in exp_unpaid if x['due_calc'] <= TODAY + timedelta(days=7))

def aging_bucket(x):
    if x['due_calc'] >= TODAY: return 0
    d = (TODAY - x['due_calc']).days
    return 1 if d <= 30 else (2 if d <= 60 else 3)

aging = defaultdict(float)
for x in inc_unpaid:
    aging[aging_bucket(x)] += x['uo']

by_client = defaultdict(float)
for x in inc_unpaid:
    by_client[x['client']] += x['uo']
top_debtors = sorted(by_client.items(), key=lambda kv: -kv[1])

by_supplier = defaultdict(float)
for x in exp_unpaid:
    by_supplier[x['supplier']] += x['rem']
top_suppliers = sorted(by_supplier.items(), key=lambda kv: -kv[1])

# Незакриті витрати по угодах: понесені витрати мінус виставлені доходи (по відкритих угодах)
inc_by_deal = defaultdict(float)
for x in inc: inc_by_deal[x['deal']] += x['uo']
exp_by_deal = defaultdict(float)
for x in exp: exp_by_deal[x['deal']] += x['uo']
exp_paid_by_deal = defaultdict(float)
for x in exp: exp_paid_by_deal[x['deal']] += x['paid_uo']
for d in deals:
    d['billed'] = inc_by_deal.get(d['num'], 0.0)
    d['spent'] = exp_by_deal.get(d['num'], 0.0)
    d['spent_paid'] = exp_paid_by_deal.get(d['num'], 0.0)
    d['unclosed'] = max(0.0, d['spent_paid'] - d['billed'])
unclosed_total = sum(d['unclosed'] for d in deals)
deals.sort(key=lambda d: -d['unclosed'])

# ── Аналіз по клієнтах: зависає vs дохід vs терміни ──
open_deal_nums = {d['num'] for d in deals}
cli = defaultdict(lambda: dict(deb=0.0, deb_age=0.0, unc=0.0, unc_age=0.0, rev=0.0, spent=0.0))
for x in inc_unpaid:
    c = client_of(x, x['client'])
    cli[c]['deb'] += x['uo']
    if x['date']: cli[c]['deb_age'] += x['uo'] * (TODAY - x['date']).days
for x in inc:
    c = client_of(x, x['client'])
    cli[c]['rev'] += x['uo']
# незакриті витрати та їх вік — по відкритих угодах
exp_age_by_deal = defaultdict(lambda: [0.0, 0.0])
for x in exp:
    ref_date = x['paid_date'] or x['date']
    if x['deal'] in open_deal_nums and ref_date and x['paid_uo'] > 0:
        exp_age_by_deal[x['deal']][0] += x['paid_uo'] * (TODAY - ref_date).days
        exp_age_by_deal[x['deal']][1] += x['paid_uo']
for d in deals:
    if d['unclosed'] > 0:
        c = all_deals.get(d['num'], {}).get('client', d['client'])
        cli[c]['unc'] += d['unclosed']
        s_age, s_uo = exp_age_by_deal.get(d['num'], (0.0, 0.0))
        avg_age = (s_age / s_uo) if s_uo else 0
        cli[c]['unc_age'] += d['unclosed'] * avg_age
# маржа по всіх угодах клієнта (доходи 01.03+ мінус витрати 01.10+ — період неповний!)
for dn, dd in all_deals.items():
    cli[dd['client']]['spent'] += exp_by_deal.get(dn, 0.0)
client_rows = []
for c, v in cli.items():
    hang = v['deb'] + v['unc']
    if hang < 0.01 and v['rev'] < 0.01: continue
    client_rows.append(dict(
        client=c, deb=v['deb'],
        deb_age=(v['deb_age'] / v['deb'] if v['deb'] else 0),
        unc=v['unc'], unc_age=(v['unc_age'] / v['unc'] if v['unc'] else 0),
        hang=hang, rev=v['rev'], margin=v['rev'] - v['spent'],
        ratio=(hang / v['rev'] if v['rev'] else None),
    ))
client_rows.sort(key=lambda z: -z['hang'])

# Кешфлоу факт: останні 4 повні тижні
cf_weeks = []
for i in range(4, 0, -1):
    ws_, we_ = MON - timedelta(days=7 * i), MON - timedelta(days=7 * i - 6)
    cin = sum(x['uo'] for x in inc if x['paid_date'] and ws_ <= x['paid_date'] <= we_)
    cout = sum(x['uo'] - x.get('rem', 0.0) for x in exp if x['paid_date'] and ws_ <= x['paid_date'] <= we_)
    cf_weeks.append((ws_, we_, cin, cout))

# Календар прогноз: 5 тижнів від поточного (прострочене → у поточний)
cal_weeks = []
run = money_uo
for i in range(5):
    ws_, we_ = MON + timedelta(days=7 * i), MON + timedelta(days=7 * i + 6)
    if i == 0:
        din = sum(x['uo'] for x in inc_unpaid if x['due_calc'] <= we_)
        dout = sum(x['rem'] for x in exp_unpaid if x['due_calc'] <= we_)
    else:
        din = sum(x['uo'] for x in inc_unpaid if ws_ <= x['due_calc'] <= we_)
        dout = sum(x['rem'] for x in exp_unpaid if ws_ <= x['due_calc'] <= we_)
    run += din - dout
    cal_weeks.append((ws_, we_, din, dout, run))

print('=== КЛЮЧОВІ ЦИФРИ (все в У.О. = USD) ===')
print(f'Гроші сьогодні: {money_uo:,.2f}')
print(f'Дебіторка: {debitorka:,.2f} ({len(inc_unpaid)} рах.), прострочено: {sum(x["uo"] for x in overdue_in):,.2f} ({len(overdue_in)} рах.)')
print(f'Aging: в строк {aging[0]:,.2f} | 1-30: {aging[1]:,.2f} | 31-60: {aging[2]:,.2f} | 60+: {aging[3]:,.2f}')
print(f'Кредиторка: {kreditorka:,.2f} ({len(exp_unpaid)} рах.), прострочено: {sum(x["rem"] for x in overdue_ex):,.2f}, аванси: {avansy:,.2f}')
print(f'Зарезервовано 7 днів: {res7:,.2f} | Вільні: {money_uo - res7:,.2f}')
print('Топ боржників:', [(k, round(v)) for k, v in top_debtors[:5]])
print('DSO по топ-клієнтах:', {k: dso.get(k) for k, _ in top_debtors[:5]})
print('Кешфлоу тижні:', [(str(w[0]), round(w[2]), round(w[3])) for w in cf_weeks])
print('Календар тижні:', [(str(w[0]), round(w[2]), round(w[3]), round(w[4])) for w in cal_weeks])
print(f'Незакриті витрати по угодах: {unclosed_total:,.2f} ({sum(1 for d in deals if d["unclosed"]>0)} угод з {len(deals)})')
print('ПО КЛІЄНТАХ (зависає | вік деб | незакр | дохід | зависає/дохід):')
for z in client_rows[:8]:
    ratio = f"{z['ratio']*100:.0f}%" if z['ratio'] is not None else '—'
    print(f"  {z['client'][:24]:24s} {z['hang']:>9,.0f} | {z['deb_age']:>3.0f} дн | {z['unc']:>8,.0f} | {z['rev']:>9,.0f} | {ratio}")
print(f'Касовий розрив 30дн: {sum(w[2] for w in cal_weeks) - sum(w[3] for w in cal_weeks):,.2f}')
EOF_ANALYSIS = True

# ═══════════ ЗБІРКА ФАЙЛУ ═══════════
HDR_FILL = PatternFill('solid', start_color='2A78D6')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BASE = Font(name='Arial', size=10)
BOLD = Font(name='Arial', bold=True, size=10)
BIG = Font(name='Arial', bold=True, size=14)
TITLE = Font(name='Arial', bold=True, size=12, color='1C5CAB')
NOTE = Font(name='Arial', size=9, italic=True, color='898781')
DETAIL = Font(name='Arial', size=9, color='52514E')
DETAIL_HDR = Font(name='Arial', size=9, bold=True, color='52514E')
LINK = Font(name='Arial', size=10, color='1C5CAB', underline='single')
CAL_HDR = Font(name='Arial', size=9, bold=True, color='52514E')
NUM = '#,##0'
NUM2 = '#,##0.00'
DATE_F = 'DD.MM.YYYY'
THIN_TOP = Border(top=Side(style='thin'))
RED = 'D03B3B'; GREEN = '006300'

def style_table(ws, ncols, nrows, num_cols=(), date_cols=()):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for r in range(2, 2 + nrows):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BASE
            if c in num_cols: cell.number_format = NUM2
            if c in date_cols: cell.number_format = DATE_F

wb = Workbook()

# ── Дебіторка (31 неоплачений доходний) ──
ws = wb.create_sheet('Дебіторка')
ws.append(['Номер', 'Дата', 'Клієнт', 'Угода', 'Контейнери', 'Валюта', 'Сума (вал.)',
           'Сума У.О.', 'Планова (Експ.)', 'Очікувана оплата', 'Днів прострочено', 'Статус', 'Вид оплати'])
inc_unpaid.sort(key=lambda x: x['due_calc'])
for i, x in enumerate(inc_unpaid, start=2):
    ws.append([x['num'], x['date'], x['client'], x['deal'], x['containers'][:30], x['cur'],
               x['amount'], x['uo'], x['due'], x['due_calc'],
               f'=MAX(0,TODAY()-J{i})', f'=IF(K{i}>0,"Прострочено","В строк")', x['kasa']])
n_deb = len(inc_unpaid)
style_table(ws, 13, n_deb, num_cols=(7, 8), date_cols=(2, 9, 10))
for col, w in zip('ABCDEFGHIJKL', (8, 11, 24, 8, 16, 8, 13, 12, 13, 13, 12, 12)):
    ws.column_dimensions[col].width = w
ws.conditional_formatting.add(f'L2:L{n_deb + 1}',
    CellIsRule(operator='equal', formula=['"Прострочено"'], font=Font(color=RED, bold=True)))
ws[f'A{n_deb + 3}'] = ('«Очікувана оплата» = дата рахунку + середній фактичний термін оплати клієнта '
                       '(планова дата з Експедитора ненадійна). Прострочення і прогноз — від очікуваної')
ws[f'A{n_deb + 3}'].font = NOTE
DEB_LAST = n_deb + 1

# ── Кредиторка (5 неоплачених витратних) ──
ws = wb.create_sheet('Кредиторка')
ws.append(['Номер', 'Дата', 'Постачальник', 'Стаття', 'Аванс', 'Валюта', 'Сума У.О.',
           'Залишок У.О.', 'Планова (Експ.)', 'Очікувана оплата', 'Примітка', 'Вид оплати'])
exp_unpaid.sort(key=lambda x: x['due_calc'])
for x in exp_unpaid:
    ws.append([x['num'], x['date'], x['supplier'], x['item'], 'Так' if x['advance'] else 'Ні',
               x['cur'], x['uo'], x['rem'], x['due'], x['due_calc'], x['note'][:60], x['kasa']])
n_kred = len(exp_unpaid)
style_table(ws, 12, n_kred, num_cols=(7, 8), date_cols=(2, 9, 10))
for col, w in zip('ABCDEFGHIJK', (8, 11, 24, 22, 7, 8, 12, 12, 13, 13, 40)):
    ws.column_dimensions[col].width = w
KRED_LAST = n_kred + 1

# ── Гроші ──
ws = wb.create_sheet('Гроші')
ws.append(['Каса / рахунок', 'Валюта', 'Залишок (вал.)', 'Залишок У.О.'])
for b in bal:
    ws.append([b['acc'], b['cur'], b['amount'], b['uo']])
ws.append(['РАЗОМ У.О.', '', '', f'=SUM(D2:D{len(bal) + 1})'])
style_table(ws, 4, len(bal) + 1, num_cols=(3, 4))
r_tot = len(bal) + 2
for c in 'AD':
    ws[f'{c}{r_tot}'].font = BOLD; ws[f'{c}{r_tot}'].border = THIN_TOP
for col, w in zip('ABCD', (30, 9, 15, 15)):
    ws.column_dimensions[col].width = w
BAL_LAST = len(bal) + 1

# ── Угоди (незакриті) ──
ws = wb.create_sheet('Угоди')
ws.append(['Номер', 'Дата', 'Клієнт', 'Статус', 'Тип', 'Маршрут', 'Вантаж',
           'Виставлено доходів У.О.', 'Оплачено витрат У.О.', 'Заморожено (оплачено − виставлено) У.О.'])
for d in deals:
    ws.append([d['num'], d['date'], d['client'], d['status'], d['typ'], d['route'],
               d['cargo'][:24], round(d['billed'], 2), round(d['spent_paid'], 2), round(d['unclosed'], 2)])
style_table(ws, 10, len(deals), num_cols=(8, 9, 10), date_cols=(2,))
for col, w in zip('ABCDEFGHIJ', (8, 11, 22, 17, 18, 30, 14, 15, 15, 15)):
    ws.column_dimensions[col].width = w
n_deals = len(deals)
ws[f'A{n_deals + 3}'] = ('Заморожено = ОПЛАЧЕНІ витрати по угоді мінус виставлені клієнту доходи. '
                         'Неоплачені рахунки постачальників сюди не входять — вони в Кредиторці')
ws[f'A{n_deals + 3}'].font = NOTE
DEALS_LAST = n_deals + 1

# ── Клієнти ──
ws = wb.create_sheet('Клієнти')
ws.append(['Клієнт', 'Дебіторка У.О.', 'Вік дебіторки, дн', 'Незакриті витрати У.О.', 'Вік витрат, дн',
           'РАЗОМ ЗАВИСАЄ У.О.', 'Виставлено доходів У.О. (з 01.03)', 'Маржа по угодах У.О.', 'Зависає / Дохід'])
for z in client_rows:
    ws.append([z['client'], round(z['deb'], 2), round(z['deb_age']), round(z['unc'], 2), round(z['unc_age']),
               round(z['hang'], 2), round(z['rev'], 2), round(z['margin'], 2),
               (round(z['ratio'], 2) if z['ratio'] is not None else '—')])
n_cli = len(client_rows)
ws.append(['РАЗОМ', f'=SUM(B2:B{n_cli + 1})', '', f'=SUM(D2:D{n_cli + 1})', '',
           f'=SUM(F2:F{n_cli + 1})', f'=SUM(G2:G{n_cli + 1})', f'=SUM(H2:H{n_cli + 1})', ''])
style_table(ws, 9, n_cli, num_cols=(2, 4, 6, 7, 8))
for c in 'ABDFGH':
    ws[f'{c}{n_cli + 2}'].font = BOLD; ws[f'{c}{n_cli + 2}'].border = THIN_TOP
    if c != 'A': ws[f'{c}{n_cli + 2}'].number_format = NUM2
ws[f'I2'].number_format = '0%'
for rr in range(2, n_cli + 2):
    ws[f'I{rr}'].number_format = '0%'
ws.conditional_formatting.add(f'H2:H{n_cli + 1}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True)))
for col, w in zip('ABCDEFGHI', (24, 14, 13, 16, 12, 16, 18, 15, 12)):
    ws.column_dimensions[col].width = w
ws[f'A{n_cli + 4}'] = ('Клієнт визначається через угоду рахунку; вік дебіторки — від дати виставлення клієнту, вік витрат — від ДАТИ ОПЛАТИ постачальнику, зважено за сумою. '
                       'Маржа = доходи (з 01.03) мінус витрати (з 01.10) по угодах клієнта — по старих угодах доходи можуть бути поза періодом')
ws[f'A{n_cli + 4}'].font = NOTE

# ── План оплат ──
ws = wb.create_sheet('План оплат')
ws.append(['Оплатити до', 'Постачальник', 'Рахунок №', 'Стаття', 'Сума У.О.', 'Тиждень', 'Мій коментар'])
for i in range(n_kred):
    r = i + 2
    ws.append([f'=Кредиторка!J{r}', f'=Кредиторка!C{r}', f'=Кредиторка!A{r}', f'=Кредиторка!D{r}',
               f'=Кредиторка!H{r}', f'="Тиждень "&ISOWEEKNUM(MAX(A{r},TODAY()))', ''])
ws.append(['', 'РАЗОМ', '', '', f'=SUM(E2:E{n_kred + 1})', '', ''])
style_table(ws, 7, n_kred, num_cols=(5,), date_cols=(1,))
for c in ('B' + str(n_kred + 2), 'E' + str(n_kred + 2)):
    ws[c].font = BOLD; ws[c].border = THIN_TOP
ws[f'E{n_kred + 2}'].number_format = NUM2
for col, w in zip('ABCDEFG', (13, 24, 10, 22, 14, 12, 28)):
    ws.column_dimensions[col].width = w

# ── Рух грошей (факт) ──
ws = wb.create_sheet('Рух грошей')
ws.append(['Тиждень', 'З', 'По', 'Надходження факт У.О.', 'Виплати факт У.О.', 'Чистий потік У.О.'])
for i, (ws_, we_, cin, cout) in enumerate(cf_weeks):
    r = i + 2
    ws.append([f'Тиждень {ws_.isocalendar()[1]} ({ws_:%d.%m}–{we_:%d.%m})', ws_, we_,
               round(cin, 2), round(cout, 2), f'=D{r}-E{r}'])
ws.append(['РАЗОМ', '', '', '=SUM(D2:D5)', '=SUM(E2:E5)', '=D6-E6'])
style_table(ws, 6, 5, num_cols=(4, 5, 6), date_cols=(2, 3))
for c in 'ADEF':
    ws[f'{c}6'].font = BOLD; ws[f'{c}6'].border = THIN_TOP
    if c != 'A': ws[f'{c}6'].number_format = NUM2
for col, w in zip('ABCDEF', (24, 11, 11, 20, 18, 17)):
    ws.column_dimensions[col].width = w

# ── Календар (прогноз, формулами від сьогодні) ──
ws = wb.create_sheet('Календар')
ws.append(['Тиждень', 'З', 'По', 'Прогноз надходжень У.О.', 'Планові оплати У.О.',
           'Чистий потік У.О.', 'Прогнозний залишок У.О.'])
for i in range(5):
    r = i + 2
    start = '=TODAY()-WEEKDAY(TODAY(),3)' if i == 0 else f'=C{r - 1}+1'
    lo = '' if i == 0 else f'Дебіторка!$J$2:$J${DEB_LAST},">="&B{r},'
    lo_k = '' if i == 0 else f'Кредиторка!$I$2:$I${KRED_LAST},">="&B{r},'
    ws.append([
        f'="Тиждень "&ISOWEEKNUM(B{r})&" ("&TEXT(B{r},"DD.MM")&"–"&TEXT(C{r},"DD.MM")&")"',
        start, f'=B{r}+6',
        f'=SUMIFS(Дебіторка!$H$2:$H${DEB_LAST},{lo}Дебіторка!$J$2:$J${DEB_LAST},"<="&C{r})',
        f'=SUMIFS(Кредиторка!$H$2:$H${KRED_LAST},{lo_k}Кредиторка!$J$2:$J${KRED_LAST},"<="&C{r})',
        f'=D{r}-E{r}',
        f'=Дашборд!$C$18+SUM($F$2:F{r})',
    ])
style_table(ws, 7, 5, num_cols=(4, 5, 6, 7), date_cols=(2, 3))
for col, w in zip('ABCDEFG', (24, 11, 11, 22, 20, 17, 22)):
    ws.column_dimensions[col].width = w
ws.conditional_formatting.add('G2:G6',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True)))
ws['A8'] = 'Поточний тиждень включає прострочене (припущення: заплатять найближчим часом). Прогноз — за плановими датами оплати з Експедитора'
ws['A8'].font = NOTE

# ── Знімки ──
ws = wb.create_sheet('Знімки')
ws.append(['Дата', 'Заморожено У.О.', 'Гроші У.О.', 'Дебіторка У.О.', 'Кредиторка У.О.'])
ws.append([TODAY, round(debitorka + avansy + unclosed_total, 2), round(money_uo, 2), round(debitorka, 2), round(kreditorka, 2)])
style_table(ws, 5, 1, num_cols=(2, 3, 4, 5), date_cols=(1,))
for col, w in zip('ABCDE', (12, 16, 14, 16, 16)):
    ws.column_dimensions[col].width = w
ws['A4'] = 'Скрипт щоранку дописуватиме рядок — з нього рахуватиметься динаміка за тиждень і спарклайни'
ws['A4'].font = NOTE

# ── Дашборд ──
ws = wb['Sheet']; ws.title = 'Дашборд'
ws['A1'] = 'ОПЕРАТИВНИЙ ФІНАНСОВИЙ ЗВІТ'; ws['A1'].font = Font(name='Arial', bold=True, size=16)
ws['D1'] = 'СТАНОМ НА 03.07.2026 (дані вивантажено вручну)'
ws['D1'].font = Font(name='Arial', bold=True, size=12, color=RED)
ws['A2'] = 'Всі суми — в У.О. (USD-еквівалент Експедитора) · гроші аналізуються ПО КАСАХ: перекидання між касами не передбачається'
ws['A2'].font = NOTE

def put(cl, cv, label, formula, bold=False, big=False, link=None, det=False, det_hdr=False, fmt=NUM):
    c = ws[cl]; c.value = label
    if det_hdr: c.font = DETAIL_HDR
    elif det: c.font = DETAIL; c.alignment = Alignment(indent=2)
    elif link: c.font = LINK; c.hyperlink = link
    else: c.font = BOLD if bold else BASE
    if formula is not None:
        v = ws[cv]; v.value = formula
        v.font = BIG if big else (BOLD if bold else (DETAIL if det else BASE))
        v.number_format = fmt

def L(row, *a, **k): put(f'A{row}', f'B{row}', *a, **k)
def R(row, *a, **k): put(f'D{row}', f'E{row}', *a, **k)

def title(cell, text):
    ws[cell] = text; ws[cell].font = TITLE

# ── Смуга 1 зліва: Залишки по касах ──
title('A4', '1. ЗАЛИШКИ ПО КАСАХ І РАХУНКАХ — 03.07.2026')
ws['B4'] = '(у валюті)'; ws['B4'].font = CAL_HDR; ws['B4'].alignment = Alignment(horizontal='right')
ws['C4'] = 'У.О.'; ws['C4'].font = CAL_HDR; ws['C4'].alignment = Alignment(horizontal='right')
for i in range(len(bal)):
    r = 5 + i; sr = 2 + i
    ws[f'A{r}'] = f'=Гроші!A{sr}&"  ("&Гроші!B{sr}&")"'; ws[f'A{r}'].font = BASE
    ws[f'B{r}'] = f'=Гроші!C{sr}'; ws[f'B{r}'].font = BASE; ws[f'B{r}'].number_format = NUM2
    ws[f'C{r}'] = f'=Гроші!D{sr}'; ws[f'C{r}'].font = BASE; ws[f'C{r}'].number_format = NUM2
r_tot = 5 + len(bal)
ws[f'A{r_tot}'] = 'РАЗОМ У.О.'; ws[f'A{r_tot}'].font = BOLD; ws[f'A{r_tot}'].border = THIN_TOP
for c in 'BC': ws[f'{c}{r_tot}'].border = THIN_TOP
ws[f'C{r_tot}'] = f'=SUM(C5:C{r_tot - 1})'; ws[f'C{r_tot}'].font = BIG; ws[f'C{r_tot}'].number_format = NUM

# ── Смуга 1 справа: Ліквідність по касах ──
title('D4', '2. ЛІКВІДНІСТЬ ПО КАСАХ (У.О.)')
for col, txt in zip('EFGH', ('Залишок', 'До сплати', 'Очік. надходж.', 'Баланс')):
    ws[f'{col}5'] = txt; ws[f'{col}5'].font = CAL_HDR
    ws[f'{col}5'].alignment = Alignment(horizontal='right', wrap_text=True)
liq_rows = []
for b in bal:
    k = b['acc']
    o = sum(x['rem'] for x in exp_unpaid if x['kasa'] == k)
    inn = sum(x['uo'] for x in inc_unpaid if x['kasa'] == k)
    if b['uo'] or o or inn:
        liq_rows.append((k, b['uo'], o, inn, b['uo'] - o))
r = 6
for k, bl_, o, inn, net in liq_rows:
    ws[f'D{r}'] = k; ws[f'D{r}'].font = BASE
    for col, val in zip('EFGH', (bl_, o, inn, net)):
        ws[f'{col}{r}'] = round(val, 2); ws[f'{col}{r}'].font = BASE; ws[f'{col}{r}'].number_format = NUM
    r += 1
ws[f'D{r}'] = 'РАЗОМ'; ws[f'D{r}'].font = BOLD; ws[f'D{r}'].border = THIN_TOP
for col, idx in zip('EFGH', (1, 2, 3, 4)):
    ws[f'{col}{r}'] = round(sum(x[idx] for x in liq_rows), 2)
    ws[f'{col}{r}'].font = BOLD; ws[f'{col}{r}'].number_format = NUM; ws[f'{col}{r}'].border = THIN_TOP
ws.conditional_formatting.add(f'H6:H{r}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True)))
LIQ_LAST = r
ws[f'D{r + 1}'] = 'Баланс = залишок каси − неоплачені рахунки, призначені до сплати з цієї каси (за «Видом оплати»)'
ws[f'D{r + 1}'].font = NOTE

# ── Смуга 1 справа нижче: Кешфлоу ──
CF0 = LIQ_LAST + 3
title(f'D{CF0}', '3. КЕШФЛОУ — факт, останні 4 тижні (У.О.)')
for col, txt in zip('EFG', ('Надходження факт', 'Виплати факт', 'Чистий потік')):
    ws[f'{col}{CF0 + 1}'] = txt; ws[f'{col}{CF0 + 1}'].font = CAL_HDR
    ws[f'{col}{CF0 + 1}'].alignment = Alignment(horizontal='right', wrap_text=True)
for i in range(4):
    r = CF0 + 2 + i; sr = 2 + i
    ws[f'D{r}'] = f"='Рух грошей'!A{sr}"; ws[f'D{r}'].font = BASE
    for col, src in zip('EFG', 'DEF'):
        ws[f'{col}{r}'] = f"='Рух грошей'!{src}{sr}"
        ws[f'{col}{r}'].font = BASE; ws[f'{col}{r}'].number_format = NUM
CF_TOT = CF0 + 6
ws[f'D{CF_TOT}'] = 'РАЗОМ'; ws[f'D{CF_TOT}'].font = BOLD; ws[f'D{CF_TOT}'].border = THIN_TOP
for col in 'EFG':
    ws[f'{col}{CF_TOT}'].border = THIN_TOP; ws[f'{col}{CF_TOT}'].font = BOLD; ws[f'{col}{CF_TOT}'].number_format = NUM
ws[f'E{CF_TOT}'] = f'=SUM(E{CF0 + 2}:E{CF0 + 5})'
ws[f'F{CF_TOT}'] = f'=SUM(F{CF0 + 2}:F{CF0 + 5})'
ws[f'G{CF_TOT}'] = f'=E{CF_TOT}-F{CF_TOT}'

# ── Смуга 2: Заморожений капітал | Дебіторка + Оплати постачальникам ──
B2 = max(r_tot, CF_TOT) + 3
title(f'A{B2}', '4. ЗАМОРОЖЕНИЙ КАПІТАЛ (У.О.)')
L(B2 + 1, 'Дебіторська заборгованість ▸', f'=SUM(Дебіторка!H2:H{DEB_LAST})', link="#'Дебіторка'!A1")
L(B2 + 2, 'Аванси постачальникам ▸', f'=SUMIFS(Кредиторка!H2:H{KRED_LAST},Кредиторка!E2:E{KRED_LAST},"Так")',
  link="#'Кредиторка'!A1")
L(B2 + 3, 'Незакриті витрати по угодах ▸', f'=SUM(Угоди!J2:J{DEALS_LAST})', link="#'Угоди'!A1")
L(B2 + 4, 'РАЗОМ ЗАМОРОЖЕНО', f'=B{B2 + 1}+B{B2 + 2}+B{B2 + 3}', bold=True, big=True)
L(B2 + 5, 'Зміна за тиждень', f'=IFERROR(B{B2 + 4}-INDEX(Знімки!B:B,MATCH(TODAY()-7,Знімки!A:A,0)),"—")')
ws[f'B{B2 + 5}'].number_format = '+#,##0;-#,##0;0'
ws.conditional_formatting.add(f'B{B2 + 5}', CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=RED, bold=True)))
ws.conditional_formatting.add(f'B{B2 + 5}', CellIsRule(operator='lessThan', formula=['0'], font=Font(color=GREEN, bold=True)))
ws[f'A{B2 + 7}'] = 'Топ угод із незакритими витратами:'; ws[f'A{B2 + 7}'].font = DETAIL_HDR
for i, d in enumerate([d for d in deals if d['unclosed'] > 0][:5]):
    r = B2 + 8 + i
    ws[f'A{r}'] = f"{d['num']} · {d['client']}"; ws[f'A{r}'].font = DETAIL
    ws[f'A{r}'].alignment = Alignment(indent=2)
    ws[f'B{r}'] = round(d['unclosed'], 2); ws[f'B{r}'].font = DETAIL; ws[f'B{r}'].number_format = NUM

title(f'D{B2}', '5. ДЕБІТОРКА (У.О.)')
R(B2 + 1, 'РАЗОМ ДЕБІТОРКА ▸', f'=SUM(Дебіторка!H2:H{DEB_LAST})', bold=True, big=True, link="#'Дебіторка'!A1")
R(B2 + 2, 'В строк', f'=SUMIFS(Дебіторка!H2:H{DEB_LAST},Дебіторка!J2:J{DEB_LAST},">="&TODAY())')
R(B2 + 3, 'Прострочено 1–30 днів', f'=SUMIFS(Дебіторка!H2:H{DEB_LAST},Дебіторка!J2:J{DEB_LAST},"<"&TODAY(),Дебіторка!J2:J{DEB_LAST},">="&TODAY()-30)')
R(B2 + 4, 'Прострочено 31–60 днів', f'=SUMIFS(Дебіторка!H2:H{DEB_LAST},Дебіторка!J2:J{DEB_LAST},"<"&TODAY()-30,Дебіторка!J2:J{DEB_LAST},">="&TODAY()-60)')
R(B2 + 5, 'Прострочено 60+ днів', f'=SUMIFS(Дебіторка!H2:H{DEB_LAST},Дебіторка!J2:J{DEB_LAST},"<"&TODAY()-60)')
for rr in (B2 + 3, B2 + 4, B2 + 5):
    ws.conditional_formatting.add(f'E{rr}', CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=RED, bold=True)))
ws[f'D{B2 + 7}'] = 'Топ боржників:'; ws[f'D{B2 + 7}'].font = DETAIL_HDR
ws[f'F{B2 + 7}'] = 'платить у сер. за'; ws[f'F{B2 + 7}'].font = CAL_HDR
ws[f'F{B2 + 7}'].alignment = Alignment(horizontal='right')
for i, (cl, s) in enumerate(top_debtors[:6]):
    r = B2 + 8 + i
    R(r, cl, f'=SUMIF(Дебіторка!C$2:C${DEB_LAST},"{cl}",Дебіторка!H$2:H${DEB_LAST})', det=True)
    d = dso.get(cl)
    ws[f'F{r}'] = f'{d} дн.' if d else 'н/д'
    ws[f'F{r}'].font = DETAIL; ws[f'F{r}'].alignment = Alignment(horizontal='right')

SUP0 = B2 + 15
title(f'D{SUP0}', '6. ОПЛАТИ ПОСТАЧАЛЬНИКАМ (У.О.)')
R(SUP0 + 1, 'РАЗОМ ДО СПЛАТИ ▸', f'=SUM(Кредиторка!H2:H{KRED_LAST})', bold=True, big=True, link="#'План оплат'!A1")
R(SUP0 + 2, 'з них прострочено (розрах.)', f'=SUMIFS(Кредиторка!H2:H{KRED_LAST},Кредиторка!J2:J{KRED_LAST},"<"&TODAY())')
ws.conditional_formatting.add(f'E{SUP0 + 2}', CellIsRule(operator='greaterThan', formula=['0'], font=Font(color=RED, bold=True)))
R(SUP0 + 3, 'до сплати найближчі 7 днів', f'=SUMIFS(Кредиторка!H2:H{KRED_LAST},Кредиторка!J2:J{KRED_LAST},"<="&TODAY()+7)')
ws[f'D{SUP0 + 4}'] = 'Топ постачальників:'; ws[f'D{SUP0 + 4}'].font = DETAIL_HDR
for i, (sp, s) in enumerate(top_suppliers[:5]):
    R(SUP0 + 5 + i, sp, f'=SUMIF(Кредиторка!C$2:C${KRED_LAST},"{sp}",Кредиторка!H$2:H${KRED_LAST})', det=True)

# ── Смуга 3: Платіжний календар | Діаграма ──
CAL0 = max(B2 + 13, SUP0 + 10) + 2
title(f'A{CAL0}', '="7. ПЛАТІЖНИЙ КАЛЕНДАР — прогноз до "&TEXT(Календар!C6,"DD.MM.YYYY")&" (У.О.)"')
for col, txt in zip('BCD', ('Прогноз надходжень', 'Планові оплати', 'Прогнозний залишок')):
    ws[f'{col}{CAL0 + 1}'] = txt; ws[f'{col}{CAL0 + 1}'].font = CAL_HDR
    ws[f'{col}{CAL0 + 1}'].alignment = Alignment(horizontal='right', wrap_text=True)
for i in range(5):
    r = CAL0 + 2 + i; sr = 2 + i
    ws[f'A{r}'] = f'=Календар!A{sr}'; ws[f'A{r}'].font = BASE
    for col, src in zip('BCD', 'DEG'):
        ws[f'{col}{r}'] = f'=Календар!{src}{sr}'
        ws[f'{col}{r}'].font = BASE; ws[f'{col}{r}'].number_format = NUM
CAL_T = CAL0 + 7
ws[f'A{CAL_T}'] = 'РАЗОМ'; ws[f'A{CAL_T}'].font = BOLD; ws[f'A{CAL_T}'].border = THIN_TOP
for col in 'BCD': ws[f'{col}{CAL_T}'].border = THIN_TOP
ws[f'B{CAL_T}'] = f'=SUM(B{CAL0 + 2}:B{CAL0 + 6})'; ws[f'B{CAL_T}'].font = BOLD; ws[f'B{CAL_T}'].number_format = NUM
ws[f'C{CAL_T}'] = f'=SUM(C{CAL0 + 2}:C{CAL0 + 6})'; ws[f'C{CAL_T}'].font = BOLD; ws[f'C{CAL_T}'].number_format = NUM
ws[f'A{CAL_T + 1}'] = 'КАСОВИЙ РОЗРИВ'; ws[f'A{CAL_T + 1}'].font = BOLD
ws[f'B{CAL_T + 1}'] = f'=B{CAL_T}-C{CAL_T}'; ws[f'B{CAL_T + 1}'].font = BIG
ws[f'B{CAL_T + 1}'].number_format = '+#,##0;-#,##0;0'
put(f'A{CAL_T + 2}', f'B{CAL_T + 2}', 'Оплати по постачальниках — аркуш «План оплат» ▸', None, link="#'План оплат'!A1")
ws.conditional_formatting.add(f'B{CAL_T + 1}', CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True, size=14)))
ws.conditional_formatting.add(f'B{CAL_T + 1}', CellIsRule(operator='greaterThanOrEqual', formula=['0'], font=Font(color=GREEN, bold=True, size=14)))
ws.conditional_formatting.add(f'D{CAL0 + 2}:D{CAL0 + 6}', CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True)))

ws[f'A{CAL_T + 4}'] = ('Сині назви ▸ — аркуші деталізації. Прострочення/прогноз — від розрахункових дат (дата рахунку + середній факт. термін). '
                       'Ліквідність по касах: значення на момент вивантаження; в автоматичній версії оновлюються щоранку. '
                       'УВАГА: прогнозний залишок у календарі — сукупний по всіх касах; реальна доступність грошей — у блоці 2.')
ws[f'A{CAL_T + 4}'].font = NOTE
for col, w in zip('ABCDEFGH', (36, 15, 13, 30, 12, 12, 13, 12)):
    ws.column_dimensions[col].width = w
ws.conditional_formatting.add(f'G{CF0 + 2}:G{CF_TOT}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color=RED, bold=True)))


cal = wb['Календар']
ch = BarChart(); ch.type = 'col'; ch.title = 'Прогноз надходжень vs планові оплати (У.О.)'
data = Reference(cal, min_col=4, max_col=5, min_row=1, max_row=6)
cats = Reference(cal, min_col=1, min_row=2, max_row=6)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ch.height, ch.width = 9, 14
ws.add_chart(ch, f'F{CAL0}')

order = ['Дашборд', 'Клієнти', 'Календар', 'Рух грошей', 'План оплат', 'Дебіторка', 'Кредиторка', 'Угоди', 'Гроші', 'Знімки']
wb._sheets = [wb[n] for n in order]
out = 'Оперативний_фінансовий_звіт_03.07.2026.xlsx'
wb.save(out)
print('SAVED', out)
