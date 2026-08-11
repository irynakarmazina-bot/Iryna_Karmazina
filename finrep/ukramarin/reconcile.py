#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Звірка «Баланс з Украмарином» (файл користувачки) ↔ каса «Каса USD Украмарин» в Експедиторі.

ДЖЕРЕЛА
  * баланс   — xlsx користувачки, аркуш Sheet1: дата | дохід | коментар | витрата | коментар
  * каса     — finrep/ukramarin/expeditor_um_cash.csv, знято з
               НАПРЯМУ з 1С (OData, регістр «Хозрасчетный» + субконто ВидОплаты), а НЕ з
               normalized/cash_moves.csv — той файл кнопка «Підтягнути свіжі дані» не оновлює
               (станом на 11.08.2026 він від 23.07 і в ньому бракує і хвоста, і приходу 1 386,00).
               Звірено з «Випискою банку» 1С: вікно 02.02–11.08.2026 дає прихід 117 706,00 /
               витрати 118 309,00 і кінцеве сальдо −176,60 — збіг до копійки.

ЩО РОБИТЬ
  1. Порівнює підсумки по кожній стороні.
  2. Зіставляє рядки по сумі всередині своєї сторони (дохід↔надходження, витрата↔витрата),
     спершу точні збіги, потім «один рядок балансу = кілька рухів каси» (баланс агрегує).
  3. Показує, що лишилося незіставленим з кожного боку — це і є розбіжності.

НІЧОГО НЕ ЗМІНЮЄ І НЕ ВИДАЛЯЄ — тільки читає і друкує.

Запуск: python3 finrep/ukramarin/reconcile.py <шлях-до-xlsx>
"""
import csv
import itertools
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
CASH = os.path.join(HERE, "expeditor_um_cash.csv")
# Кінець покритого періоду. Дані знято напряму з 1С (OData) станом на 11.08.2026.
CASH_LAST_DATE = "2026-08-11"
DATE_TOL = 90          # днів: баланс і Експедитор ставлять різні дати одній операції
DATE_WARN = 30         # більший розрив дат — показати окремо як розбіжність дати
MAX_COMBO = 4          # скільки рухів каси максимум складаємо в один рядок балансу


def money(v):
    return "{:,.2f}".format(v).replace(",", " ").replace(".", ",")


def days(a, b):
    import datetime
    fa = datetime.date(*map(int, a.split("-")))
    fb = datetime.date(*map(int, b.split("-")))
    return abs((fa - fb).days)


# ------------------------------------------------------------------ баланс (xlsx)
def read_balance(path):
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)["Sheet1"]
    out, last = [], None
    for r in range(2, 52):                       # рядок 54 — підсумки =SUM(B2:B51)
        d = ws.cell(row=r, column=1).value
        if d is not None:
            # У файлі жовтень–грудень проставлені 2026 роком, хоча йдуть ПЕРЕД січнем 2026
            # і збігаються з рухами каси за жовтень–грудень 2025 (див. звіт).
            y = d.year - 1 if d.month >= 10 and d.year == 2026 else d.year
            last = "%04d-%02d-%02d" % (y, d.month, d.day)
        inc = ws.cell(row=r, column=2).value
        exp = ws.cell(row=r, column=4).value
        note = " ".join(str(ws.cell(row=r, column=c).value or "").strip() for c in (3, 5)).strip()
        if isinstance(inc, (int, float)) and inc:
            out.append({"row": r, "date": last, "side": "in", "amt": round(float(inc), 2), "note": note})
        if isinstance(exp, (int, float)) and exp:
            out.append({"row": r, "date": last, "side": "out", "amt": round(float(exp), 2), "note": note})
    return out


# ------------------------------------------------------------------ каса (Експедитор)
def read_cash():
    out = []
    with open(CASH, encoding="utf-8") as f:
        for i, r in enumerate(csv.DictReader(f), start=2):
            inc, exp = float(r["income"] or 0), float(r["expense"] or 0)
            note = " ".join(x for x in (r["counterparty"], r["note"]) if x).strip()
            if r["transfer"]:
                note = ("[переміщення] " + note).strip()
            if inc:
                out.append({"row": i, "date": r["date"], "side": "in", "amt": round(inc, 2), "note": note})
            if exp:
                out.append({"row": i, "date": r["date"], "side": "out", "amt": round(exp, 2), "note": note})
    return out


# ------------------------------------------------------------------ зіставлення
def match(bal, cash):
    """Спершу 1:1 по точній сумі, потім 1:N (баланс агрегує кілька рухів каси)."""
    pairs, free_b, free_c = [], list(bal), list(cash)

    for b in list(free_b):                                   # 1:1
        best = None
        for c in free_c:
            if c["side"] == b["side"] and abs(c["amt"] - b["amt"]) < 0.005 and days(b["date"], c["date"]) <= DATE_TOL:
                d = days(b["date"], c["date"])
                if best is None or d < best[0]:
                    best = (d, c)
        if best:
            pairs.append((b, [best[1]], "1:1"))
            free_b.remove(b)
            free_c.remove(best[1])

    for b in list(free_b):                                   # 1:N — баланс агрегує кілька рухів
        pool = [c for c in free_c if c["side"] == b["side"] and days(b["date"], c["date"]) <= DATE_TOL]
        hit = None
        for n in range(2, MAX_COMBO + 1):
            for combo in itertools.combinations(pool, n):
                if abs(sum(c["amt"] for c in combo) - b["amt"]) < 0.005:
                    hit = list(combo)
                    break
            if hit:
                break
        if hit:
            pairs.append((b, hit, "1:%d" % len(hit)))
            free_b.remove(b)
            for c in hit:
                free_c.remove(c)

    for c in list(free_c):                                   # N:1 — баланс дробить один рух каси
        pool = [b for b in free_b if b["side"] == c["side"] and days(b["date"], c["date"]) <= DATE_TOL]
        hit = None
        for n in range(2, MAX_COMBO + 1):
            for combo in itertools.combinations(pool, n):
                if abs(sum(x["amt"] for x in combo) - c["amt"]) < 0.005:
                    hit = list(combo)
                    break
            if hit:
                break
        if hit:
            pairs.append((hit, [c], "%d:1" % len(hit)))
            free_c.remove(c)
            for b in hit:
                free_b.remove(b)
    return pairs, free_b, free_c


def block(title, rows, src):
    print("\n" + title)
    if not rows:
        print("   — немає")
        return 0.0
    tot = 0.0
    for r in sorted(rows, key=lambda x: (x["date"], -x["amt"])):
        tot += r["amt"]
        print("   %s  %-12s  %-3s  %s  (%s %d)"
              % (r["date"], money(r["amt"]), "дох" if r["side"] == "in" else "вит",
                 (r["note"] or "—")[:52], src, r["row"]))
    print("   РАЗОМ: %s" % money(tot))
    return tot


def main():
    path = sys.argv[1]
    bal, cash = read_balance(path), read_cash()

    def tot(rows, side, upto=None):
        return round(sum(r["amt"] for r in rows
                         if r["side"] == side and (upto is None or r["date"] <= upto)), 2)

    print("=" * 78)
    print("ЗВІРКА: баланс з Украмарином  ↔  каса «Каса USD Украмарин» (Експедитор)")
    print("=" * 78)
    print("\nПІДСУМКИ ЗА ВЕСЬ ФАЙЛ")
    print("   баланс:      дохід %s   витрата %s   сальдо %s"
          % (money(tot(bal, "in")), money(tot(bal, "out")), money(tot(bal, "in") - tot(bal, "out"))))
    print("   Експедитор:  надх. %s   витрата %s   сальдо %s"
          % (money(tot(cash, "in")), money(tot(cash, "out")), money(tot(cash, "in") - tot(cash, "out"))))
    print("\nПІДСУМКИ ДО %s (де є з чим порівнювати)" % CASH_LAST_DATE)
    bi, bo = tot(bal, "in", CASH_LAST_DATE), tot(bal, "out", CASH_LAST_DATE)
    ci, co = tot(cash, "in", CASH_LAST_DATE), tot(cash, "out", CASH_LAST_DATE)
    print("   баланс:      дохід %s   витрата %s   сальдо %s" % (money(bi), money(bo), money(bi - bo)))
    print("   Експедитор:  надх. %s   витрата %s   сальдо %s" % (money(ci), money(co), money(ci - co)))
    print("   РІЗНИЦЯ:     дохід %s   витрата %s   сальдо %s"
          % (money(bi - ci), money(bo - co), money((bi - bo) - (ci - co))))

    pairs, free_b, free_c = match(bal, cash)
    agg = [p for p in pairs if p[2] != "1:1"]
    print("\nЗІСТАВИЛОСЬ: %d рядків балансу (%d з них — один рядок балансу проти кількох рухів каси)"
          % (len(pairs), len(agg)))
    for left, right, kind in sorted(agg, key=lambda p: (p[0] if isinstance(p[0], list) else [p[0]])[0]["date"]):
        ls = left if isinstance(left, list) else [left]
        print("   %s %s: %s  =  %s   «%s»"
              % (ls[0]["date"], "дох" if ls[0]["side"] == "in" else "вит",
                 " + ".join(money(x["amt"]) for x in ls),
                 " + ".join(money(c["amt"]) for c in right),
                 (ls[0]["note"] or "—")[:34]))

    far = []
    for left, right, kind in pairs:
        ls = left if isinstance(left, list) else [left]
        gap = max(days(x["date"], c["date"]) for x in ls for c in right)
        if gap > DATE_WARN:
            far.append((gap, ls, right))
    print("\n⚠️ ЗІЙШЛОСЬ ПО СУМІ, АЛЕ РІЗНІ ДАТИ (розрив > %d днів)" % DATE_WARN)
    if not far:
        print("   — немає")
    for gap, ls, right in sorted(far, key=lambda x: -x[0]):
        print("   %s  %s %s «%s»  →  в Експедиторі %s  (розрив %d дн.)"
              % (ls[0]["date"], money(sum(x["amt"] for x in ls)),
                 "дох" if ls[0]["side"] == "in" else "вит", (ls[0]["note"] or "—")[:34],
                 ", ".join(c["date"] for c in right), gap))

    a = block("❗ Є В БАЛАНСІ, НЕМАЄ В КАСІ ЕКСПЕДИТОРА", free_b, "рядок")
    b_ = block("❗ Є В КАСІ ЕКСПЕДИТОРА, НЕМАЄ В БАЛАНСІ",
               [c for c in free_c if c["date"] <= CASH_LAST_DATE], "рух")
    print("\nСАЛЬДО НЕЗІСТАВЛЕНОГО: %s (баланс) − %s (Експедитор)" % (money(a), money(b_)))


if __name__ == "__main__":
    main()
